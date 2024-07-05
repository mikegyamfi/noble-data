import calendar
import hashlib
import hmac
import json
from datetime import datetime

from decouple import config
from django.contrib.auth.forms import PasswordResetForm
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
import requests
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.http import urlsafe_base64_encode
from django.views.decorators.csrf import csrf_exempt

from . import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from . import helper, models
from .helper import generate_unique_exttrid
from .models import Payment


# Create your views here.
def home(request):
    if models.Announcement.objects.filter(active=True).exists():
        announcement = models.Announcement.objects.filter(active=True).first()
        messages.info(request, announcement.message)
        return render(request, "layouts/index.html")
    return render(request, "layouts/index.html")


def services(request):
    return render(request, "layouts/services.html")


from django.shortcuts import render, redirect
from django.http import JsonResponse
from . import models
from . import helper
import requests


def pay_with_wallet(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.first().phone_number  # Updated query
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = float(request.POST.get("amount"))  # Selling price
        reference = request.POST.get("reference")

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < amount:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})

        if user.status == "User":
            bundle = models.IshareBundlePrice.objects.get(price=amount).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentIshareBundlePrice.objects.get(price=amount).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentIshareBundlePrice.objects.get(price=amount).bundle_volume
        else:
            bundle = models.IshareBundlePrice.objects.get(price=amount).bundle_volume

        send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, reference)
        data = send_bundle_response.json()
        print(data)

        sms_headers = {
            'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            'Content-Type': 'application/json'
        }
        sms_url = 'https://webapp.usmsgh.com/api/sms/send'

        if send_bundle_response.status_code == 200:
            if data["code"] == "0000":
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Completed"
                )
                new_transaction.save()
                user.wallet -= amount  # Subtract selling price from wallet balance
                user.save()

                receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {reference}\n"
                sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nCurrent Wallet Balance: {user.wallet}\nThank you for using Noble Data."

                response1 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{request.user.phone}&from=Noble Data&sms={sms_message}")
                print(response1.text)

                response2 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to={phone_number}&from=Noble Data&sms={receiver_message}")
                print(response2.text)

                if user.status == "User":
                    bundle = models.IshareBundlePrice.objects.get(price=amount)
                elif user.status == "Agent":
                    bundle = models.AgentIshareBundlePrice.objects.get(price=amount)
                elif user.status == "Super Agent":
                    bundle = models.SuperAgentIshareBundlePrice.objects.get(price=amount)
                else:
                    bundle = models.IshareBundlePrice.objects.get(price=amount)

                # Get the purchase price based on the selling price
                bundle_price_obj = bundle
                if not bundle_price_obj:
                    return JsonResponse({'status': 'Invalid amount'})
                purchase_price = bundle_price_obj.purchase_price

                profit = amount - purchase_price

                # Create ProfitInstance
                new_profit_instance = models.ProfitInstance.objects.create(
                    selling_price_total=amount,
                    purchase_price_total=purchase_price,
                    profit=profit,
                    channel="AT",  # Set your channel here based on user status
                )
                new_profit_instance.save()

                new_wallet_transaction = models.WalletTransaction.objects.create(
                    user=request.user,
                    transaction_type="Debit",
                    transaction_amount=float(amount),
                    transaction_use="AT",
                    new_balance=user.wallet
                )
                new_wallet_transaction.save()

                return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
            else:
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=request.user,
                    bundle_number=phone_number,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Failed"
                )
                new_transaction.save()
                return JsonResponse({'status': 'Something went wrong'})
    return redirect('airtel-tigo')


@login_required(login_url='login')
def airtel_tigo(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    db_user_id = request.user.id
    status = user.status
    form = forms.IShareBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email
    if request.method == "POST":
        if models.AdminInfo.objects.filter().first().active_payment == "AppsNMobile":
            payment_reference = helper.generate_unique_exttrid()
            amount_paid = request.POST.get("offers")
            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=payment_reference,
                amount=amount_paid,
                transaction_date=datetime.now(),
                transaction_status="Pending"
            )
            new_payment.save()
            phone_number = request.POST.get("phone_number")
            offer = request.POST.get("offers")
            print(offer)
            if user.status == "User":
                bundle = models.IshareBundlePrice.objects.get(price=offer).bundle_volume
            elif user.status == "Agent":
                bundle = models.AgentIshareBundlePrice.objects.get(price=offer).bundle_volume
            elif user.status == "Super Agent":
                bundle = models.SuperAgentIshareBundlePrice.objects.get(price=offer).bundle_volume
            else:
                bundle = models.IshareBundlePrice.objects.get(price=offer).bundle_volume

            new_payment.transaction_details = {
                "amount": offer,
                "bundle": bundle,
                "phone_number": phone_number,
                "payment_reference": payment_reference,
                "channel": "ishare"
            }
            new_payment.save()
            return redirect("anmgw_payment_checkout", reference=new_payment.reference)
    #     form = forms.IShareBundleForm(data=request.POST, status=status)
    #     if form.is_valid():
    #         phone_number = form.cleaned_data["phone_number"]
    #         amount = form.cleaned_data["offers"]
    #
    #         print(amount.price)
    #
    #         details = {
    #             'phone_number': phone_number,
    #             'offers': amount.price
    #         }
    #
    #         new_payment = models.Payment.objects.create(
    #             user=request.user,
    #             reference=reference,
    #             transaction_date=datetime.now(),
    #             transaction_details=details,
    #             channel="ishare",
    #         )
    #         new_payment.save()
    #         print("payment saved")
    #         print("form valid")
    #
    #         url = "https://payproxyapi.hubtel.com/items/initiate"
    #
    #         payload = json.dumps({
    #             "totalAmount": amount.price,
    #             "description": "Payment for AT Bundle",
    #             "callbackUrl": "https://www.nobledatagh.com/hubtel_webhook",
    #             "returnUrl": "https://www.nobledatagh.com",
    #             "cancellationUrl": "https://www.nobledatagh.com",
    #             "merchantAccountNumber": "2019630",
    #             "clientReference": new_payment.reference
    #         })
    #         headers = {
    #             'Content-Type': 'application/json',
    #             'Authorization': 'Basic T0VWRTlCRTphYTVhNDc3YTI3M2Q0NWViODlkZTk4YThmMWYzZDQwMw=='
    #         }
    #
    #         response = requests.request("POST", url, headers=headers, data=payload)
    #
    #         data = response.json()
    #
    #         checkoutUrl = data['data']['checkoutUrl']
    #
    #         return redirect(checkoutUrl)
    # if request.method == "POST":
    #     form = forms.IShareBundleForm(data=request.POST, status=status)
    #     payment_reference = request.POST.get("reference")
    #     amount_paid = request.POST.get("amount")
    #     new_payment = models.Payment.objects.create(
    #         user=request.user,
    #         reference=payment_reference,
    #         amount=amount_paid,
    #         transaction_date=datetime.now(),
    #         transaction_status="Completed"
    #     )
    #     new_payment.save()
    #     print("payment saved")
    #     print("form valid")
    #     phone_number = request.POST.get("phone")
    #     offer = request.POST.get("amount")
    #     print(offer)
    #     bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume if user.status == "User" else models.AgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
    #     new_transaction = models.IShareBundleTransaction.objects.create(
    #         user=request.user,
    #         bundle_number=phone_number,
    #         offer=f"{bundle}MB",
    #         reference=payment_reference,
    #         transaction_status="Pending"
    #     )
    #     print("created")
    #     new_transaction.save()
    #
    #     print("===========================")
    #     print(phone_number)
    #     print(bundle)
    #     send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, payment_reference)
    #     data = send_bundle_response.json()
    #
    #     print(data)
    #
    #     sms_headers = {
    #         'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
    #         'Content-Type': 'application/json'
    #     }
    #
    #     sms_url = 'https://webapp.usmsgh.com/api/sms/send'
    #
    #     if send_bundle_response.status_code == 200:
    #         if data["code"] == "0000":
    #             transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
    #             print("got here")
    #             print(transaction_to_be_updated.transaction_status)
    #             transaction_to_be_updated.transaction_status = "Completed"
    #             transaction_to_be_updated.save()
    #             print(request.user.phone)
    #             print("***********")
    #             receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {request.user.phone}.\nReference: {payment_reference}\n"
    #             sms_message = f"Hello @{request.user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {payment_reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"
    #
    #             num_without_0 = phone_number[1:]
    #             print(num_without_0)
    #             receiver_body = {
    #                 'recipient': f"233{num_without_0}",
    #                 'sender_id': 'Noble Data',
    #                 'message': receiver_message
    #             }
    #
    #             response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
    #             print(response.text)
    #
    #             sms_body = {
    #                 'recipient': f"233{request.user.phone}",
    #                 'sender_id': 'Noble Data',
    #                 'message': sms_message
    #             }
    #
    #             response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #
    #             print(response.text)
    #
    #             return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
    #         else:
    #             transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
    #             transaction_to_be_updated.transaction_status = "Failed"
    #             new_transaction.save()
    #             sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"
    #
    #             sms_body = {
    #                 'recipient': f"233{request.user.phone}",
    #                 'sender_id': 'Noble Data',
    #                 'message': sms_message
    #             }
    #             response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #             print(response.text)
    #             # r_sms_url = f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=UmpEc1JzeFV4cERKTWxUWktqZEs&to={phone_number}&from=Noble Data GH&sms={receiver_message}"
    #             # response = requests.request("GET", url=r_sms_url)
    #             # print(response.text)
    #             return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    #     else:
    #         transaction_to_be_updated = models.IShareBundleTransaction.objects.get(reference=payment_reference)
    #         transaction_to_be_updated.transaction_status = "Failed"
    #         new_transaction.save()
    #         sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {payment_reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"
    #
    #         sms_body = {
    #             'recipient': f'233{request.user.phone}',
    #             'sender_id': 'Noble Data',
    #             'message': sms_message
    #         }
    #
    #         response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #
    #         print(response.text)
    #         return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
    user = models.CustomUser.objects.get(id=request.user.id)
    payment_method = models.AdminInfo.objects.filter().first().active_payment
    context = {"form": form, 'id': db_user_id, "ref": reference, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet, "payment_method": payment_method}
    return render(request, "layouts/services/at.html", context=context)


def mtn_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = float(request.POST.get("amount"))  # Selling price
        reference = request.POST.get("reference")

        admin = models.AdminInfo.objects.first().phone_number
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < amount:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})

        # Get the purchase price from the MTNBundlePrice model
        # bundle_price_obj = models.MTNBundlePrice.objects.get(price=amount)
        # purchase_price = bundle_price_obj.purchase_price
        #
        # # Calculate profit based on selling price and purchase price
        # profit = amount - purchase_price

        # Get the bundle volume based on user status
        if user.status == "User":
            bundle = models.MTNBundlePrice.objects.get(price=amount).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentMTNBundlePrice.objects.get(price=amount).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentMTNBundlePrice.objects.get(price=amount).bundle_volume
        else:
            bundle = models.MTNBundlePrice.objects.get(price=amount).bundle_volume

        # Assume some logic to initiate the transaction
        # url = "https://posapi.bestpaygh.com/api/v1/initiate_mtn_transaction"
        # payload = { ... }
        # headers = { ... }
        # response = requests.request("POST", url, headers=headers, data=json.dumps(payload))
        # print(response.text)

        sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
        new_mtn_transaction = models.MTNTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= amount
        user.save()
        if user.status == "User":
            bundle = models.MTNBundlePrice.objects.get(price=amount)
        elif user.status == "Agent":
            bundle = models.AgentMTNBundlePrice.objects.get(price=amount)
        elif user.status == "Super Agent":
            bundle = models.SuperAgentMTNBundlePrice.objects.get(price=amount)
        else:
            bundle = models.MTNBundlePrice.objects.get(price=amount)

        bundle_price_obj = bundle
        if not bundle_price_obj:
            return JsonResponse({'status': 'Invalid amount'})
        purchase_price = bundle_price_obj.purchase_price

        profit = amount - purchase_price

        # Create ProfitInstance
        new_profit_instance = models.ProfitInstance.objects.create(
            selling_price_total=amount,
            purchase_price_total=purchase_price,
            profit=profit,
            channel="MTN",  # Set your channel here based on user status
        )
        new_profit_instance.save()

        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=request.user,
            transaction_type="Debit",
            transaction_amount=float(amount),
            transaction_use="MTN",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()

        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('mtn')


@login_required(login_url='login')
def big_time_pay_with_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = float(request.POST.get("amount"))  # Selling price
        reference = request.POST.get("reference")

        admin = models.AdminInfo.objects.first().phone_number
        if user.wallet is None:
            return JsonResponse({'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < amount:
            return JsonResponse({'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        # Get the purchase price from the BigTimeBundlePrice model
        # bundle_price_obj = models.BigTimeBundlePrice.objects.get(price=amount)
        # purchase_price = bundle_price_obj.purchase_price
        #
        # # Calculate profit based on selling price and purchase price
        # profit = amount - purchase_price

        # Get the bundle volume based on user status
        if user.status == "User":
            bundle = models.BigTimeBundlePrice.objects.get(price=amount).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentBigTimeBundlePrice.objects.get(price=amount).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=amount).bundle_volume
        else:
            bundle = models.BigTimeBundlePrice.objects.get(price=amount).bundle_volume

        # Create BigTimeTransaction
        new_big_time_transaction = models.BigTimeTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_big_time_transaction.save()

        user.wallet -= amount
        user.save()

        if user.status == "User":
            bundle = models.BigTimeBundlePrice.objects.get(price=amount)
        elif user.status == "Agent":
            bundle = models.AgentBigTimeBundlePrice.objects.get(price=amount)
        elif user.status == "Super Agent":
            bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=amount)
        else:
            bundle = models.BigTimeBundlePrice.objects.get(price=amount)

        bundle_price_obj = bundle
        if not bundle_price_obj:
            return JsonResponse({'status': 'Invalid amount'})
        purchase_price = bundle_price_obj.purchase_price

        profit = amount - purchase_price

        # Create ProfitInstance
        new_profit_instance = models.ProfitInstance.objects.create(
            selling_price_total=amount,
            purchase_price_total=purchase_price,
            profit=profit,
            channel="BigTime",  # Set your channel here based on user status
        )
        new_profit_instance.save()

        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=request.user,
            transaction_type="Debit",
            transaction_amount=float(amount),
            transaction_use="Big Time",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()

        # Send SMS
        sms_headers = {
            'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            'Content-Type': 'application/json'
        }
        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello,\nA Big time order has been placed.\nReference: {reference}.\nThank you"

        sms_body = {
            'recipient': f"233549914001",
            'sender_id': 'Noble Data',
            'message': sms_message
        }

        try:
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            print("Could not send message")
            pass

        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})

    return redirect('big_time')


@login_required(login_url='login')
def mtn(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    db_user_id = request.user.id
    phone = user.phone
    status = user.status
    form = forms.MTNForm(status=status)
    reference = helper.ref_generator()
    user_email = request.user.email
    if request.method == "POST":
        if models.AdminInfo.objects.filter().first().active_payment == "AppsNMobile":
            payment_reference = helper.generate_unique_exttrid()
            amount_paid = request.POST.get("offers")
            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=payment_reference,
                amount=amount_paid,
                transaction_date=datetime.now(),
                transaction_status="Pending"
            )
            new_payment.save()
            phone_number = request.POST.get("phone_number")
            offer = request.POST.get("offers")
            print(offer)
            if user.status == "User":
                bundle = models.MTNBundlePrice.objects.get(price=offer).bundle_volume
            elif user.status == "Agent":
                bundle = models.AgentMTNBundlePrice.objects.get(price=offer).bundle_volume
            elif user.status == "Super Agent":
                bundle = models.SuperAgentMTNBundlePrice.objects.get(price=offer).bundle_volume
            else:
                bundle = models.MTNBundlePrice.objects.get(price=offer).bundle_volume

            new_payment.transaction_details = {
                "amount": offer,
                "bundle": bundle,
                "phone_number": phone_number,
                "payment_reference": payment_reference,
                "channel": "mtn"
            }
            new_payment.save()
            return redirect("anmgw_payment_checkout", reference=new_payment.reference)
    # if request.method == "POST":
    #     form = forms.MTNForm(data=request.POST, status=status)
    #     if form.is_valid():
    #         phone_number = form.cleaned_data['phone_number']
    #         amount = form.cleaned_data['offers']
    #
    #         print(amount.price)
    #
    #         details = {
    #             'phone_number': f"0{phone_number}",
    #             'offers': amount.price
    #         }
    #
    #         new_payment = models.Payment.objects.create(
    #             user=request.user,
    #             reference=reference,
    #             transaction_date=datetime.now(),
    #             transaction_details=details,
    #             channel="mtn",
    #         )
    #         new_payment.save()
    #         print("payment saved")
    #         print("form valid")
    #
    #         url = "https://payproxyapi.hubtel.com/items/initiate"
    #
    #         payload = json.dumps({
    #             "totalAmount": amount.price,
    #             "description": "Payment for MTN Bundle",
    #             "callbackUrl": "https://www.nobledatagh.com/hubtel_webhook",
    #             "returnUrl": "https://www.nobledatagh.com",
    #             "cancellationUrl": "https://www.nobledatagh.com",
    #             "merchantAccountNumber": "2019630",
    #             "clientReference": new_payment.reference
    #         })
    #         headers = {
    #             'Content-Type': 'application/json',
    #             'Authorization': 'Basic T0VWRTlCRTphYTVhNDc3YTI3M2Q0NWViODlkZTk4YThmMWYzZDQwMw=='
    #         }
    #
    #         response = requests.request("POST", url, headers=headers, data=payload)
    #
    #         data = response.json()
    #
    #         checkoutUrl = data['data']['checkoutUrl']
    #
    #         return redirect(checkoutUrl)
    # if request.method == "POST":
    #     payment_reference = request.POST.get("reference")
    #     amount_paid = request.POST.get("amount")
    #     new_payment = models.Payment.objects.create(
    #         user=request.user,
    #         reference=payment_reference,
    #         amount=amount_paid,
    #         transaction_date=datetime.now(),
    #         transaction_status="Completed"
    #     )
    #     new_payment.save()
    #     phone_number = request.POST.get("phone")
    #     offer = request.POST.get("amount")
    #
    #     bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume if user.status == "User" else models.AgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
    #     url = "https://posapi.bestpaygh.com/api/v1/initiate_mtn_transaction"
    #
    #     payload = json.dumps({
    #         "user_id": user_id,
    #         "receiver": phone_number,
    #         "data_volume": bundle,
    #         "reference": reference,
    #         "amount": offer,
    #         "channel": phone
    #     })
    #     headers = {
    #         'Authorization': auth,
    #         'Content-Type': 'application/json'
    #     }
    #
    #     response = requests.request("POST", url, headers=headers, data=payload)
    #
    #     print(response.text)
    #     print(phone_number)
    #     new_mtn_transaction = models.MTNTransaction.objects.create(
    #         user=request.user,
    #         bundle_number=phone_number,
    #         offer=f"{bundle}MB",
    #         reference=payment_reference,
    #     )
    #     new_mtn_transaction.save()
    #     # sms_headers = {
    #     #     'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
    #     #     'Content-Type': 'application/json'
    #     # }
    #     #
    #     # sms_url = 'https://webapp.usmsgh.com/api/sms/send'
    #     # sms_message = f"An order has been placed. {bundle}MB for {phone_number}"
    #     # admin = models.AdminInfo.objects.filter().first().phone_number
    #     # sms_body = {
    #     #     'recipient': f"233{admin}",
    #     #     'sender_id': 'Noble Data',
    #     #     'message': sms_message
    #     # }
    #     # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
    #     # print(response.text)
    #     return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    phone_num = user.phone
    mtn_dict = {}

    if user.status == "Agent":
        mtn_offer = models.AgentMTNBundlePrice.objects.all()
    else:
        mtn_offer = models.MTNBundlePrice.objects.all()
    for offer in mtn_offer:
        mtn_dict[str(offer)] = offer.bundle_volume
    payment_method = models.AdminInfo.objects.filter().first().active_payment
    context = {'form': form, 'phone_num': phone_num, 'id': db_user_id,
               'mtn_dict': json.dumps(mtn_dict), "ref": reference, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet, "payment_method": payment_method}
    return render(request, "layouts/services/mtn.html", context=context)


@login_required(login_url='login')
def afa_registration(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    reference = helper.ref_generator()
    price = models.AdminInfo.objects.filter().first().afa_price
    user_email = request.user.email
    print(price)
    # if request.method == "POST":
    #     form = forms.AFARegistrationForm(request.POST)
    #     if form.is_valid():
    #         # name = transaction_details["name"]
    #         # phone_number = transaction_details["phone"]
    #         # gh_card_number = transaction_details["card"]
    #         # occupation = transaction_details["occupation"]
    #         # date_of_birth = transaction_details["date_of_birth"]
    #         details = {
    #             "name": form.cleaned_data["name"],
    #             "phone": form.cleaned_data["phone_number"],
    #             "card": form.cleaned_data["gh_card_number"],
    #             "occupation": form.cleaned_data["occupation"],
    #             "date_of_birth": form.cleaned_data["date_of_birth"],
    #             "region": form.cleaned_data["region"]
    #         }
    #         new_payment = models.Payment.objects.create(
    #             user=request.user,
    #             reference=reference,
    #             transaction_details=details,
    #             transaction_date=datetime.now(),
    #             channel="afa"
    #         )
    #         new_payment.save()
    #
    #         url = "https://payproxyapi.hubtel.com/items/initiate"
    #
    #         payload = json.dumps({
    #             "totalAmount": price,
    #             "description": "Payment for AFA Registration",
    #             "callbackUrl": "https://www.nobledatagh.com/hubtel_webhook",
    #             "returnUrl": "https://www.nobledatagh.com",
    #             "cancellationUrl": "https://www.nobledatagh.com",
    #             "merchantAccountNumber": "2019630",
    #             "clientReference": new_payment.reference
    #         })
    #         headers = {
    #             'Content-Type': 'application/json',
    #             'Authorization': 'Basic T0VWRTlCRTphYTVhNDc3YTI3M2Q0NWViODlkZTk4YThmMWYzZDQwMw=='
    #         }
    #
    #         response = requests.request("POST", url, headers=headers, data=payload)
    #
    #         data = response.json()
    #
    #         checkoutUrl = data['data']['checkoutUrl']
    #
    #         return redirect(checkoutUrl)
    form = forms.AFARegistrationForm()
    context = {'form': form, 'ref': reference, 'price': price, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/afa.html", context=context)


def afa_registration_wallet(request):
    if request.method == "POST":
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        name = request.POST.get("name")
        card_number = request.POST.get("card")
        occupation = request.POST.get("occupation")
        date_of_birth = request.POST.get("birth")
        region = request.POST.get("region")
        price = models.AdminInfo.objects.filter().first().afa_price

        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        new_registration = models.AFARegistration.objects.create(
            user=user,
            reference=reference,
            name=name,
            phone_number=phone_number,
            gh_card_number=card_number,
            occupation=occupation,
            date_of_birth=date_of_birth,
            region=region
        )
        new_registration.save()
        user.wallet -= float(price)
        user.save()

        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=request.user,
            transaction_type="Debit",
            transaction_amount=float(amount),
            transaction_use="AFA",
            new_balance=user.wallet
        )
        new_wallet_transaction.save()
        sms_headers = {
            'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello,\nAn AFA Registration order has been placed.\nReference: {reference}.\nThank you"

        sms_body = {
            'recipient': f"233549914001",
            'sender_id': 'Noble Data',
            'message': sms_message
        }
        try:
            print("tried")
            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            print(response.text)
        except:
            print("could not send message")
            pass
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('home')


@login_required(login_url='login')
def big_time(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.BigTimeBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email

    # if request.method == "POST":
    #     form = forms.BigTimeBundleForm(data=request.POST, status=status)
    #     if form.is_valid():
    #         phone_number = form.cleaned_data['phone_number']
    #         amount = form.cleaned_data['offers']
    #         details = {
    #             'phone_number': phone_number,
    #             'offers': amount.price
    #         }
    #         new_payment = models.Payment.objects.create(
    #             user=request.user,
    #             reference=reference,
    #             transaction_details=details,
    #             transaction_date=datetime.now(),
    #             channel="bigtime"
    #         )
    #         new_payment.save()
    #
    #         url = "https://payproxyapi.hubtel.com/items/initiate"
    #
    #         payload = json.dumps({
    #             "totalAmount": amount.price,
    #             "description": "Payment for AFA Registration",
    #             "callbackUrl": "https://www.nobledatagh.com/hubtel_webhook",
    #             "returnUrl": "https://www.nobledatagh.com",
    #             "cancellationUrl": "https://www.nobledatagh.com",
    #             "merchantAccountNumber": "2019630",
    #             "clientReference": new_payment.reference
    #         })
    #         headers = {
    #             'Content-Type': 'application/json',
    #             'Authorization': 'Basic T0VWRTlCRTphYTVhNDc3YTI3M2Q0NWViODlkZTk4YThmMWYzZDQwMw=='
    #         }
    #
    #         response = requests.request("POST", url, headers=headers, data=payload)
    #
    #         data = response.json()
    #
    #         checkoutUrl = data['data']['checkoutUrl']
    #
    #         return redirect(checkoutUrl)
    user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    context = {'form': form,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/big_time.html", context=context)


@csrf_exempt
def paystack_webhook(request):
    if request.method == "POST":
        paystack_secret_key = config("PAYSTACK_SECRET_KEY")
        # print(paystack_secret_key)
        payload = json.loads(request.body)

        paystack_signature = request.headers.get("X-Paystack-Signature")

        if not paystack_secret_key or not paystack_signature:
            return HttpResponse(status=400)

        computed_signature = hmac.new(
            paystack_secret_key.encode('utf-8'),
            request.body,
            hashlib.sha512
        ).hexdigest()

        if computed_signature == paystack_signature:
            print("yes")
            print(payload.get('data'))
            r_data = payload.get('data')
            print(r_data.get('metadata'))
            print(payload.get('event'))
            if payload.get('event') == 'charge.success':
                metadata = r_data.get('metadata')
                receiver = metadata.get('receiver')
                db_id = metadata.get('db_id')
                print(db_id)
                offer = metadata.get('offer')
                user = models.CustomUser.objects.get(id=int(db_id))
                print(user)
                channel = metadata.get('channel')
                real_amount = metadata.get('real_amount')
                print(real_amount)
                paid_amount = r_data.get('amount')
                reference = r_data.get('reference')

                slashed_amount = (float(paid_amount) / 100) * (1 - 0.0195)
                reference = r_data.get('reference')

                rounded_real_amount = round(float(real_amount))
                rounded_paid_amount = round(float(slashed_amount))

                print(f"reeeeeeeaaaaaaaaal amount: {rounded_real_amount}")
                print(f"paaaaaaaaaaaaaiiddd amount: {rounded_paid_amount}")

                is_within_range = (rounded_real_amount - 3) <= rounded_paid_amount <= (rounded_real_amount + 3)

                if not is_within_range:
                    sms_headers = {
                        'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    sms_message = f"Malicious attempt on webhook. Real amount: {rounded_real_amount} | Paid amount: {rounded_paid_amount}. Referrer: {reference}"

                    sms_body = {
                        'recipient': "233242442147",
                        'sender_id': 'Noble Data',
                        'message': sms_message
                    }
                    try:
                        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        print(response.text)
                    except:
                        pass

                    print("not within range")
                    return HttpResponse(200)

                if channel == "ishare":
                    bundle = models.IshareBundlePrice.objects.get(price=float(
                        real_amount)).bundle_volume if user.status == "User" else models.AgentIshareBundlePrice.objects.get(
                        price=float(real_amount)).bundle_volume
                    if models.IShareBundleTransaction.objects.filter(reference=reference, offer=f"{bundle}MB",
                                                                     transaction_status="Completed").exists():
                        return HttpResponse(status=200)
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Pending"
                    )
                    new_transaction.save()
                    send_bundle_response = helper.send_bundle(user, receiver, bundle, reference)
                    data = send_bundle_response.json()

                    print(data)

                    sms_headers = {
                        'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'

                    if send_bundle_response.status_code == 200:
                        if data["code"] == "0000":
                            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                                reference=reference)
                            print("got here")
                            print(transaction_to_be_updated.transaction_status)
                            transaction_to_be_updated.transaction_status = "Completed"
                            transaction_to_be_updated.save()

                            num_without_0 = receiver[1:]

                            receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                            sms_message = f"Hello @{user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {receiver}.\nReference: {reference}\nThank you for using Noble Data GH."

                            response1 = requests.get(
                                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{user.phone}&from=Noble Data&sms={sms_message}")
                            print(response1.text)

                            response2 = requests.get(
                                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to={num_without_0}&from=Noble Data&sms={receiver_message}")
                            print(response2.text)

                            bundle = models.IshareBundlePrice.objects.get(price=float(
                                real_amount)) if user.status == "User" else models.AgentIshareBundlePrice.objects.get(
                                price=float(real_amount))

                            purchase_price = bundle.purchase_price

                            profit = float(real_amount) - float(purchase_price)

                            new_profit_instance = models.ProfitInstance.objects.create(
                                selling_price_total=real_amount,
                                purchase_price_total=purchase_price,
                                profit=profit,
                                channel="AT",  # Set your channel here based on user status
                            )
                            new_profit_instance.save()

                            print(user.phone)
                            print("***********")
                            receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                            sms_message = f"Hello @{user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {receiver}.\nReference: {reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"

                            print(num_without_0)
                            # receiver_body = {
                            #     'recipient': f"233{num_without_0}",
                            #     'sender_id': 'Noble Data',
                            #     'message': receiver_message
                            # }
                            #
                            # response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                            # print(response.text)
                            #
                            # sms_body = {
                            #     'recipient': f"233{user.phone}",
                            #     'sender_id': 'Noble Data',
                            #     'message': sms_message
                            # }
                            #
                            # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                            #
                            # print(response.text)

                            return HttpResponse(status=200)
                        else:
                            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                                reference=reference)
                            transaction_to_be_updated.transaction_status = "Failed"
                            new_transaction.save()
                            sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {receiver}.\nReference: {reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"

                            sms_body = {
                                'recipient': f"233{user.phone}",
                                'sender_id': 'Noble Data',
                                'message': sms_message
                            }
                            response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                            print(response.text)
                            return HttpResponse(status=500)
                    else:
                        transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                            reference=reference)
                        transaction_to_be_updated.transaction_status = "Failed"
                        new_transaction.save()
                        sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {receiver}.\nReference: {reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"

                        sms_body = {
                            'recipient': f'233{user.phone}',
                            'sender_id': 'Noble Data',
                            'message': sms_message
                        }

                        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                        print(response.text)
                        return HttpResponse(status=500)
                elif channel == "mtn":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Completed"
                    )
                    new_payment.save()

                    if user.status == "User":
                        bundle = models.MTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentMTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    else:
                        bundle = models.MTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                    new_mtn_transaction = models.MTNTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_mtn_transaction.save()

                    print(receiver)

                    if user.status == "User":
                        bundle = models.MTNBundlePrice.objects.get(price=float(real_amount))
                    elif user.status == "Agent":
                        bundle = models.AgentMTNBundlePrice.objects.get(price=float(real_amount))
                    else:
                        bundle = models.MTNBundlePrice.objects.get(price=float(real_amount))

                    purchase_price = bundle.purchase_price

                    profit = float(real_amount) - float(purchase_price)

                    new_profit_instance = models.ProfitInstance.objects.create(
                        selling_price_total=real_amount,
                        purchase_price_total=purchase_price,
                        profit=profit,
                        channel="MTN",  # Set your channel here based on user status
                    )
                    new_profit_instance.save()

                    return HttpResponse(status=200)
                elif channel == "voda":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Completed"
                    )
                    new_payment.save()

                    if user.status == "User":
                        bundle = models.VodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentVodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                    else:
                        bundle = models.VodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                    new_voda_transaction = models.VodafoneTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_voda_transaction.save()

                    sms_message = f"Telecel order has been placed. {bundle}MB for {receiver}. Reference: {reference}"
                    response1 = requests.get(
                        f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0549914001&from=Noble Data&sms={sms_message}")
                    print(response1.text)

                    print(receiver)

                    if user.status == "User":
                        bundle = models.VodaBundlePrice.objects.get(price=float(real_amount))
                    elif user.status == "Agent":
                        bundle = models.AgentVodaBundlePrice.objects.get(price=float(real_amount))
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(real_amount))
                    else:
                        bundle = models.VodaBundlePrice.objects.get(price=float(real_amount))

                    purchase_price = bundle.purchase_price

                    profit = float(real_amount) - float(purchase_price)

                    new_profit_instance = models.ProfitInstance.objects.create(
                        selling_price_total=real_amount,
                        purchase_price_total=purchase_price,
                        profit=profit,
                        channel="Voda",  # Set your channel here based on user status
                    )
                    new_profit_instance.save()

                    return HttpResponse(status=200)
                elif channel == "at_min":
                    new_payment = models.Payment.objects.create(
                        user=user,
                        reference=reference,
                        amount=paid_amount,
                        transaction_date=datetime.now(),
                        transaction_status="Pending"
                    )
                    new_payment.save()

                    if user.status == "User":
                        minutes = models.ATCreditPrice.objects.get(price=float(real_amount)).minutes
                    elif user.status == "Agent":
                        minutes = models.AgentATCreditPrice.objects.get(price=float(real_amount)).minutes
                    elif user.status == "Super Agent":
                        minutes = models.SuperAgentATCreditPrice.objects.get(price=float(real_amount)).minutes
                    else:
                        minutes = models.ATCreditPrice.objects.get(price=float(real_amount)).minutes

                    print(receiver)

                    new_mtn_transaction = models.ATMinuteTransaction.objects.create(
                        user=user,
                        bundle_number=receiver,
                        offer=f"{minutes} Minutes",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    sms_message = f"AT Minutes order has been placed. {minutes} minutes for {receiver}. Reference: {reference}"
                    response1 = requests.get(
                        f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0549914001&from=Noble Data&sms={sms_message}")
                    print(response1.text)
                    return HttpResponse(status=200)
                elif channel == "topup":
                    amount = real_amount

                    user.wallet += float(amount)
                    user.save()

                    new_topup = models.TopUpRequest.objects.create(
                        user=user,
                        reference=reference,
                        amount=amount,
                        status=True,
                    )
                    new_topup.save()

                    new_wallet_transaction = models.WalletTransaction.objects.create(
                        user=user,
                        transaction_type="Credit",
                        transaction_amount=float(amount),
                        transaction_use="Top up (Paystack)",
                        new_balance=user.wallet
                    )
                    new_wallet_transaction.save()
                    return JsonResponse({'status': "Wallet Credited"}, status=200)
                else:
                    return HttpResponse(status=200)
            else:
                return HttpResponse(status=200)
        else:
            return HttpResponse(status=401)


@login_required(login_url='login')
def history(request):
    user_transactions = models.IShareBundleTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "AirtelTigo Transactions"
    net = "tigo"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def mtn_history(request):
    user_transactions = models.MTNTransaction.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "MTN Transactions"
    net = "mtn"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def big_time_history(request):
    user_transactions = models.BigTimeTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Big Time Transactions"
    net = "bt"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def afa_history(request):
    user_transactions = models.AFARegistration.objects.filter(user=request.user).order_by('transaction_date').reverse()
    header = "AFA Registrations"
    net = "afa"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/afa_history.html", context=context)


def verify_transaction(request, reference):
    if request.method == "GET":
        response = helper.verify_paystack_transaction(reference)
        data = response.json()
        try:
            status = data["data"]["status"]
            amount = data["data"]["amount"]
            api_reference = data["data"]["reference"]
            date = data["data"]["paid_at"]
            real_amount = float(amount) / 100
            print(status)
            print(real_amount)
            print(api_reference)
            print(reference)
            print(date)
        except:
            status = data["status"]
        return JsonResponse({'status': status})


def change_excel_status(request, status, to_change_to):
    transactions = models.MTNTransaction.objects.filter(
        transaction_status=status) if to_change_to != "Completed" else models.MTNTransaction.objects.filter(
        transaction_status=status).order_by('transaction_date')[:10]
    for transaction in transactions:
        transaction.transaction_status = to_change_to
        transaction.save()
        if to_change_to == "Completed":
            transaction_number = transaction.user.phone
            sms_headers = {
                'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"Your MTN transaction has been completed. {transaction.bundle_number} has been credited with {transaction.offer}.\nTransaction Reference: {transaction.reference}"

            sms_body = {
                'recipient': f"233{transaction_number}",
                'sender_id': 'Noble Data',
                'message': sms_message
            }
            try:
                # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                # print(response.text)
                response1 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{transaction_number}&from=Noble Data&sms={sms_message}")
                print(response1.text)
            except:
                messages.success(request, f"Transaction Completed")
                return redirect('mtn_admin', status=status)
        else:
            messages.success(request, f"Status changed from {status} to {to_change_to}")
            return redirect("mtn_admin", status=status)
    messages.success(request, f"Status changed from {status} to {to_change_to}")
    return redirect("mtn_admin", status=status)


@login_required(login_url='login')
def admin_mtn_history(request, status):
    if request.user.is_staff and request.user.is_superuser:
        if request.method == "POST":
            from io import BytesIO
            from openpyxl import load_workbook
            from django.http import HttpResponse
            import datetime

            # Assuming `uploaded_file` is the Excel file uploaded by the user
            uploaded_file = request.FILES['file'] if 'file' in request.FILES else None
            if not uploaded_file:
                messages.error(request, "No excel file found")
                return redirect('mtn_admin', status=status)

            # Load the uploaded Excel file into memory
            excel_buffer = BytesIO(uploaded_file.read())
            book = load_workbook(excel_buffer)
            sheet = book.active  # Assuming the data is on the active sheet

            # Assuming we have identified the recipient and data column indices
            # Replace these with the actual indices if available
            recipient_col_index = 1  # Example index for "RECIPIENT"
            data_col_index = 2  # Example index for "DATA"

            # Query your Django model
            queryset = models.MTNTransaction.objects.filter(transaction_status="Pending")

            # Determine the starting row for updates, preserving headers and any other pre-existing content
            start_row = 2  # Assuming data starts from row 2

            for record in queryset:
                # Assuming 'bundle_number' and 'offer' fields exist in your model
                recipient_value = str(record.bundle_number)  # Ensure it's a string to preserve formatting
                data_value = record.offer  # Adjust based on actual field type
                cleaned_data_value = float(data_value.replace('MB', ''))
                data_value_gb = round(float(cleaned_data_value) / 1000, 2)

                # Find next available row (avoid overwriting non-empty rows if necessary)
                while sheet.cell(row=start_row, column=recipient_col_index).value is not None:
                    start_row += 1

                # Update cells
                sheet.cell(row=start_row, column=recipient_col_index, value=recipient_value)
                sheet.cell(row=start_row, column=data_col_index, value=data_value_gb)

                # Update the record status, if necessary
                record.transaction_status = "Processing"
                record.save()

            # Save the modified Excel file to the buffer
            excel_buffer.seek(0)  # Reset buffer position
            book.save(excel_buffer)

            # Prepare the response with the modified Excel file
            excel_buffer.seek(0)  # Reset buffer position to read the content
            response = HttpResponse(excel_buffer.getvalue(),
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(
                datetime.datetime.now().strftime("%Y-%m-%d-%H-%M-%S"))

            return response

        all_txns = models.MTNTransaction.objects.filter(transaction_status=status).order_by('-transaction_date')[:800]
        context = {'txns': all_txns, 'status': status}
        return render(request, "layouts/services/mtn_admin.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('mtn_admin', status=status)


@login_required(login_url='login')
def admin_bt_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.BigTimeTransaction.objects.filter().order_by('-transaction_date')
        context = {'txns': all_txns}
        return render(request, "layouts/services/bt_admin.html", context=context)


@login_required(login_url='login')
def admin_afa_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.AFARegistration.objects.filter().order_by('-transaction_date')
        context = {'txns': all_txns}
        return render(request, "layouts/services/afa_admin.html", context=context)


@login_required(login_url='login')
def mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.MTNTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your MTN transaction has been been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'Noble Data',
            'message': sms_message
        }
        try:
            response1 = requests.get(
                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{txn.user.phone}&from=Noble Data&sms={sms_message}")
            print(response1.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('mtn_admin', status="Pending")
        messages.success(request, f"Transaction Completed")
        return redirect('mtn_admin', status="Pending")


@login_required(login_url='login')
def bt_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.BigTimeTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AT BIG TIME transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'Noble Data',
            'message': sms_message
        }
        try:
            response1 = requests.get(
                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{txn.user.phone}&from=Noble Data&sms={sms_message}")
            print(response1.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('bt_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('bt_admin')


@login_required(login_url='login')
def afa_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.AFARegistration.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AFA Registration has been completed. {txn.phone_number} has been registered.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'Noble Data',
            'message': sms_message
        }
        response1 = requests.get(
            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{txn.user.phone}&from=Noble Data&sms={sms_message}")
        print(response1.text)
        messages.success(request, f"Transaction Completed")
        return redirect('afa_admin')


@login_required(login_url='login')
def wallet_history(request):
    user_wallet_transactions = models.WalletTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()[:1000]
    header = "Wallet Transactions"
    net = "wallet"
    context = {'txns': user_wallet_transactions, "header": header, "net": net}
    return render(request, "layouts/wallet_history.html", context=context)


def credit_user(request):
    form = forms.CreditUserForm()
    if request.user.is_superuser:
        if request.method == "POST":
            form = forms.CreditUserForm(request.POST)
            if form.is_valid():
                user = form.cleaned_data["user"]
                amount = form.cleaned_data["amount"]
                print(user)
                print(amount)
                user_needed = models.CustomUser.objects.get(username=user)
                if user_needed.wallet is None:
                    user_needed.wallet = float(amount)
                    user_needed.wallet = float(user.wallet)
                    new_wallet_transaction = models.WalletTransaction.objects.create(
                        user=user_needed,
                        transaction_type="Credit",
                        transaction_amount=float(amount),
                        transaction_use="Top up"
                    )
                else:
                    user_needed.wallet += float(amount)
                    user_needed.wallet = float(user.wallet)
                    new_wallet_transaction = models.WalletTransaction.objects.create(
                        user=request.user,
                        transaction_type="Credit",
                        transaction_amount=float(amount),
                        transaction_use="Top up"
                    )
                user_needed.save()
                new_wallet_transaction.new_balance = user_needed.wallet
                new_wallet_transaction.save()
                print(user_needed.username)
                messages.success(request, "Crediting Successful")
                return redirect('credit_user')
        context = {'form': form}
        return render(request, "layouts/services/credit.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def topup_info(request):
    # if request.method == "POST":
    #     admin = models.AdminInfo.objects.filter().first().phone_number
    #     user = models.CustomUser.objects.get(id=request.user.id)
    #     amount = request.POST.get("amount")
    #     print(amount)
    #     reference = helper.top_up_ref_generator()
    #     details = {
    #         'topup_amount': amount
    #     }
    #     new_payment = models.Payment.objects.create(
    #         user=request.user,
    #         reference=reference,
    #         transaction_details=details,
    #         transaction_date=datetime.now(),
    #         channel="topup"
    #     )
    #     new_payment.save()
    #
    #     url = "https://payproxyapi.hubtel.com/items/initiate"
    #
    #     payload = json.dumps({
    #         "totalAmount": amount,
    #         "description": "Payment for AFA Registration",
    #         "callbackUrl": "https://www.nobledatagh.com/hubtel_webhook",
    #         "returnUrl": "https://www.nobledatagh.com",
    #         "cancellationUrl": "https://www.nobledatagh.com",
    #         "merchantAccountNumber": "2019630",
    #         "clientReference": new_payment.reference
    #     })
    #     headers = {
    #         'Content-Type': 'application/json',
    #         'Authorization': 'Basic T0VWRTlCRTphYTVhNDc3YTI3M2Q0NWViODlkZTk4YThmMWYzZDQwMw=='
    #     }
    #
    #     response = requests.request("POST", url, headers=headers, data=payload)
    #
    #     data = response.json()
    #
    #     checkoutUrl = data['data']['checkoutUrl']
    #
    #     return redirect(checkoutUrl)
    if request.method == "POST":
        active_payment = models.AdminInfo.objects.filter().first().topup_active_payment
        if active_payment == "Direct":
            admin = models.AdminInfo.objects.filter().first().phone_number
            user = models.CustomUser.objects.get(id=request.user.id)
            amount = request.POST.get("amount")
            print(amount)
            reference = helper.top_up_ref_generator()
            new_topup_request = models.TopUpRequest.objects.create(
                user=request.user,
                amount=amount,
                reference=reference,
                status=False
            )
            new_topup_request.save()

            sms_headers = {
                'Authorization': 'Bearer 1136|LwSl79qyzTZ9kbcf9SpGGl1ThsY0Ujf7tcMxvPze',
                'Content-Type': 'application/json'
            }

            sms_url = 'https://webapp.usmsgh.com/api/sms/send'
            sms_message = f"A top up request has been placed.\nGHS{amount} for {user}.\nReference: {reference}"

            sms_body = {
                'recipient': f"233{admin}",
                'sender_id': 'Noble',
                'message': sms_message
            }
            # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
            # print(response.text)
            messages.success(request,
                             f"Your Request has been sent successfully. Kindly go on to pay to {admin} and use the reference stated as reference. Reference: {reference}")
            return redirect("request_successful", reference)
        elif active_payment == "AppsNMobile":
            payment_reference = helper.generate_unique_exttrid()
            user = models.CustomUser.objects.get(id=request.user.id)
            amount_paid = request.POST.get("amount")
            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=payment_reference,
                amount=amount_paid,
                transaction_date=datetime.now(),
                transaction_status="Pending"
            )
            new_payment.save()
            amount = request.POST.get("amount")

            new_payment.transaction_details = {
                "amount": amount,
                "payment_reference": payment_reference
            }
            new_payment.save()

            new_topup_request = models.TopUpRequest.objects.create(
                user=user,
                amount=amount,
                reference=payment_reference,
                status=False
            )
            new_topup_request.save()
            return redirect("anmgw_payment_checkout", reference=new_payment.reference)
    active_payment = models.AdminInfo.objects.filter().first().topup_active_payment
    user_email = request.user.email
    db_user_id = request.user.id
    reference = helper.ref_generator()
    context = {
        'payment_method': active_payment,
        "email": user_email,
        "ref": reference,
        'id': db_user_id,
    }
    return render(request, "layouts/topup-info.html", context=context)


@login_required(login_url='login')
def request_successful(request, reference):
    admin = models.AdminInfo.objects.filter().first()
    context = {
        "name": admin.name,
        "number": f"0{admin.momo_number}",
        "channel": admin.payment_channel,
        "reference": reference
    }
    return render(request, "layouts/services/request_successful.html", context=context)


def topup_list(request):
    if request.user.is_superuser:
        topup_requests = models.TopUpRequest.objects.all().order_by('date').reverse()[:1000]
        context = {
            'requests': topup_requests,
        }
        return render(request, "layouts/services/topup_list.html", context=context)
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


@login_required(login_url='login')
def credit_user_from_list(request, reference):
    if request.user.is_superuser:
        crediting = models.TopUpRequest.objects.filter(reference=reference).first()
        user = crediting.user
        custom_user = models.CustomUser.objects.get(username=user.username)
        if crediting.status:
            return redirect('topup_list')
        amount = crediting.amount
        print(user)
        print(user.phone)
        print(amount)
        custom_user.wallet += amount
        custom_user.save()
        new_wallet_transaction = models.WalletTransaction.objects.create(
            user=custom_user,
            transaction_type="Credit",
            transaction_amount=float(amount),
            transaction_use="Top up",
            new_balance=custom_user.wallet
        )
        new_wallet_transaction.save()
        crediting.status = True
        crediting.credited_at = datetime.now()
        crediting.save()
        sms_headers = {
            'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Hello,\nYour wallet has been topped up with GHS{amount}.\nReference: {reference}.\nThank you"

        sms_body = {
            'recipient': f"233{custom_user.phone}",
            'sender_id': 'Noble Data',
            'message': sms_message
        }
        try:
            print("tried")
            response1 = requests.get(
                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{custom_user.phone}&from=Noble Data&sms={sms_message}")
            print(response1.text)
        except:
            print("could not send message")
            pass

        messages.success(request, f"{user} has been credited with {amount}")
        return redirect('topup_list')


@csrf_exempt
def hubtel_webhook(request):
    if request.method == 'POST':
        print("hit the webhook")
        try:
            payload = request.body.decode('utf-8')
            print("Hubtel payment Info: ", payload)
            json_payload = json.loads(payload)
            print(json_payload)

            data = json_payload.get('Data')
            print(data)
            reference = data.get('ClientReference')
            print(reference)
            txn_status = data.get('Status')
            txn_description = data.get('Description')
            amount = data.get('Amount')
            print(txn_status, amount)

            if txn_status == 'Success':
                print("success")
                transaction_saved = models.Payment.objects.get(reference=reference, transaction_status="Unfinished")
                transaction_saved.transaction_status = "Paid"
                transaction_saved.payment_description = txn_description
                transaction_saved.amount = amount
                transaction_saved.save()
                transaction_details = transaction_saved.transaction_details
                transaction_channel = transaction_saved.channel
                user = transaction_saved.user
                # receiver = collection_saved['number']
                # bundle_volume = collection_saved['data_volume']
                # name = collection_saved['name']
                # email = collection_saved['email']
                # phone_number = collection_saved['buyer']
                # date_and_time = collection_saved['date_and_time']
                # txn_type = collection_saved['type']
                # user_id = collection_saved['uid']
                print(transaction_details, transaction_channel)

                if transaction_channel == "ishare":
                    offer = transaction_details["offers"]
                    phone_number = transaction_details["phone_number"]

                    if user.status == "User":
                        bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentIshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    else:
                        bundle = models.IshareBundlePrice.objects.get(price=float(offer)).bundle_volume
                    new_transaction = models.IShareBundleTransaction.objects.create(
                        user=user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                        transaction_status="Pending"
                    )
                    print("created")
                    new_transaction.save()

                    print("===========================")
                    print(phone_number)
                    print(bundle)
                    print(user)
                    print(reference)
                    send_bundle_response = helper.send_bundle(user, phone_number, bundle, reference)
                    data = send_bundle_response.json()

                    print(data)

                    sms_headers = {
                        'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'

                    if send_bundle_response.status_code == 200:
                        if data["code"] == "0000":
                            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                                reference=reference)
                            print("got here")
                            print(transaction_to_be_updated.transaction_status)
                            transaction_to_be_updated.transaction_status = "Completed"
                            transaction_to_be_updated.save()
                            print(user.phone)
                            print("***********")
                            receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                            sms_message = f"Hello @{user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {phone_number}.\nReference: {reference}\nThank you for using Noble Data.\n\nThe Noble Data"

                            sms_body = {
                                'recipient': f"233{user.phone}",
                                'sender_id': 'Noble Data',
                                'message': sms_message
                            }
                            try:
                                response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                                print(response.text)
                            except:
                                print("message not sent")
                                pass
                            return JsonResponse({'status': 'Transaction Completed Successfully'}, status=200)
                        else:
                            transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                                reference=reference)
                            transaction_to_be_updated.transaction_status = "Failed"
                            new_transaction.save()
                            sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {reference}\nThank you for using Noble Data.\n\nThe Noble Data"

                            sms_body = {
                                'recipient': f"233{user.phone}",
                                'sender_id': 'Data4All',
                                'message': sms_message
                            }
                            return JsonResponse({'status': 'Something went wrong'}, status=500)
                    else:
                        transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                            reference=reference)
                        transaction_to_be_updated.transaction_status = "Failed"
                        new_transaction.save()
                        sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {phone_number}.\nReference: {reference}\nThank you for using Noble Data.\n\nThe Noble Data"

                        sms_body = {
                            'recipient': f'233{user.phone}',
                            'sender_id': 'Noble Data',
                            'message': sms_message
                        }

                        # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        #
                        # print(response.text)
                        return JsonResponse({'status': 'Something went wrong', 'icon': 'error'})
                elif transaction_channel == "mtn":
                    offer = transaction_details["offers"]
                    phone_number = transaction_details["phone_number"]

                    auth = config("AT")
                    user_id = config("USER_ID")

                    if user.status == "User":
                        bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(offer)).bundle_volume
                    else:
                        bundle = models.MTNBundlePrice.objects.get(price=float(offer)).bundle_volume

                    url = "https://posapi.bestpaygh.com/api/v1/initiate_mtn_transaction"

                    payload = json.dumps({
                        "user_id": user_id,
                        "receiver": phone_number,
                        "data_volume": bundle,
                        "reference": reference,
                        "amount": offer,
                        "channel": user.phone
                    })
                    headers = {
                        'Authorization': auth,
                        'Content-Type': 'application/json'
                    }

                    # response = requests.request("POST", url, headers=headers, data=payload)
                    #
                    # print(response.text)

                    print(phone_number)

                    print(phone_number)
                    new_mtn_transaction = models.MTNTransaction.objects.create(
                        user=user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                elif transaction_channel == "bigtime":
                    offer = transaction_details["offers"]
                    phone_number = transaction_details["phone_number"]
                    if user.status == "User":
                        bundle = models.BigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Agent":
                        bundle = models.AgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    elif user.status == "Super Agent":
                        bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    else:
                        bundle = models.SuperAgentBigTimeBundlePrice.objects.get(price=float(offer)).bundle_volume
                    print(phone_number)
                    new_mtn_transaction = models.BigTimeTransaction.objects.create(
                        user=user,
                        bundle_number=phone_number,
                        offer=f"{bundle}MB",
                        reference=reference,
                    )
                    new_mtn_transaction.save()
                    sms_headers = {
                        'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    sms_message = f"Hello,\nA Big Time order has been placed.\nReference: {reference}.\nThank you"

                    sms_body = {
                        'recipient': f"233549914001",
                        'sender_id': 'Noble Data',
                        'message': sms_message
                    }
                    try:
                        print("tried")
                        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        print(response.text)
                    except:
                        print("could not send message")
                        pass
                    return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                elif transaction_channel == "afa":
                    name = transaction_details["name"]
                    phone_number = transaction_details["phone"]
                    gh_card_number = transaction_details["card"]
                    occupation = transaction_details["occupation"]
                    date_of_birth = transaction_details["date_of_birth"]
                    region = transaction_details["region"]

                    new_afa_reg = models.AFARegistration.objects.create(
                        user=user,
                        phone_number=phone_number,
                        gh_card_number=gh_card_number,
                        name=name,
                        occupation=occupation,
                        reference=reference,
                        date_of_birth=date_of_birth,
                        region=region
                    )
                    new_afa_reg.save()

                    sms_headers = {
                        'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                        'Content-Type': 'application/json'
                    }

                    sms_url = 'https://webapp.usmsgh.com/api/sms/send'
                    sms_message = f"Hello,\nAn AFA Registration order has been placed.\nReference: {reference}.\nThank you"

                    sms_body = {
                        'recipient': f"233549914001",
                        'sender_id': 'Noble Data',
                        'message': sms_message
                    }
                    try:
                        print("tried")
                        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        print(response.text)
                    except:
                        print("could not send message")
                        pass
                    return JsonResponse({'status': "Your transaction will be completed shortly"}, status=200)
                elif transaction_channel == "topup":
                    amount = transaction_details["topup_amount"]
                    amount = amount

                    user.wallet += round(float(amount))
                    user.save()

                    new_topup = models.TopUpRequest.objects.create(
                        user=user,
                        reference=reference,
                        amount=amount,
                        status=True,
                    )
                    new_topup.save()
                    return JsonResponse({'status': "Wallet Credited"}, status=200)
                else:
                    print("no type found")
                    return JsonResponse({'message': "No Type Found"}, status=500)
            else:
                print("Transaction was not Successful")
                return JsonResponse({'message': 'Transaction Failed'}, status=200)
        except Exception as e:
            print("Error Processing hubtel webhook:", str(e))
            return JsonResponse({'status': 'error'}, status=500)
    else:
        print("not post")
        return JsonResponse({'message': 'Not Found'}, status=404)


from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator


def password_reset_request(request):
    if request.method == "POST":
        password_reset_form = PasswordResetForm(request.POST)
        if password_reset_form.is_valid():
            data = password_reset_form.cleaned_data['email']
            user = models.CustomUser.objects.filter(email=data).first()
            current_user = user
            if user:
                subject = "Password Reset Requested"
                email_template_name = "password/password_reset_message.txt"
                c = {
                    "name": user.first_name,
                    "email": user.email,
                    'domain': 'www.nobledatagh.com',
                    'site_name': 'Noble Data',
                    "uid": urlsafe_base64_encode(force_bytes(user.pk)),
                    "user": user,
                    'token': default_token_generator.make_token(user),
                    'protocol': 'https',
                }
                email = render_to_string(email_template_name, c)

                sms_headers = {
                    'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                    'Content-Type': 'application/json'
                }

                sms_url = 'https://webapp.usmsgh.com/api/sms/send'

                sms_body = {
                    'recipient': f"233{user.phone}",
                    'sender_id': 'Noble Data',
                    'message': email
                }
                # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                # print(response.text)
                response1 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{user.phone}&from=Noble Data&sms={email}")
                print(response1.text)
                return redirect("/password_reset/done/")
    password_reset_form = PasswordResetForm()
    return render(request=request, template_name="password/password_reset.html",
                  context={"password_reset_form": password_reset_form})


@login_required(login_url='login')
def profit_home(request):
    if request.user.is_superuser:
        return render(request, 'layouts/profit_page.html')
    else:
        messages.error(request, "Access Denied")
        return redirect('home')


def channel_profit(request, channel):
    profit_instances = models.ProfitInstance.objects.filter(channel=channel)

    # Calculate profit breakdown by month
    today = timezone.now()
    this_year = today.year
    monthly_data = []
    total_profit = 0
    total_selling_price = 0
    total_purchase_price = 0

    for month in range(1, 13):  # Loop through each month of the year
        month_name = calendar.month_name[month]
        month_instances = profit_instances.filter(date_and_time__year=this_year, date_and_time__month=month)
        month_profit = month_instances.aggregate(total_profit=Sum('profit'))['total_profit'] or 0
        month_selling_price = month_instances.aggregate(total_selling_price=Sum('selling_price_total'))[
                                  'total_selling_price'] or 0
        month_purchase_price = month_instances.aggregate(total_purchase_price=Sum('purchase_price_total'))[
                                   'total_purchase_price'] or 0

        monthly_data.append({
            'month': month_name,
            'profit': month_profit,
            'selling_price': month_selling_price,
            'purchase_price': month_purchase_price,
        })

        total_profit += month_profit
        total_selling_price += month_selling_price
        total_purchase_price += month_purchase_price

    context = {
        'monthly_data': monthly_data,
        'total_profit': total_profit,
        'total_selling_price': total_selling_price,
        'total_purchase_price': total_purchase_price,
        'channel': channel,
    }

    return render(request, 'layouts/services/profit.html', context)


def query_txn(request):
    if request.method == "POST":
        reference = request.POST.get('reference')
        print(reference)

        headers = {
            "api-key": config("API_KEY"),
            "api-secret": config("API_SECRET"),
        }
        response = requests.post(
            url=f"https://console.bestpaygh.com/api/flexi/v1/transaction_detail/{reference.strip()}/", headers=headers)
        data = response.json()
        print(data)
        try:
            print(data["message"])
            messages.info(request, data["message"])
        except:
            print(data['api_response']['message'])
            messages.info(request, data['api_response']['message'])

        return redirect('query_txn')
    return render(request, "layouts/query_txn.html")


@login_required(login_url='login')
def voda(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.VodaBundleForm(status)
    reference = helper.ref_generator()
    user_email = request.user.email
    db_user_id = request.user.id

    if request.method == "POST":
        if models.AdminInfo.objects.filter().first().active_payment == "AppsNMobile":
            payment_reference = helper.generate_unique_exttrid()
            amount_paid = request.POST.get("offers")
            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=payment_reference,
                amount=amount_paid,
                transaction_date=datetime.now(),
                transaction_status="Pending"
            )
            new_payment.save()
            phone_number = request.POST.get("phone_number")
            offer = request.POST.get("offers")
            print(offer)
            if user.status == "User":
                bundle = models.VodaBundlePrice.objects.get(price=offer).bundle_volume
            elif user.status == "Agent":
                bundle = models.AgentVodaBundlePrice.objects.get(price=offer).bundle_volume
            elif user.status == "Super Agent":
                bundle = models.SuperAgentVodaBundlePrice.objects.get(price=offer).bundle_volume
            else:
                bundle = models.VodaBundlePrice.objects.get(price=offer).bundle_volume

            new_payment.transaction_details = {
                "amount": offer,
                "bundle": bundle,
                "phone_number": phone_number,
                "payment_reference": payment_reference
            }
            new_payment.save()
            return redirect("anmgw_payment_checkout", reference=new_payment.reference)
        elif models.AdminInfo.objects.filter().first().active_payment == "Direct":
            payment_reference = request.POST.get("POST")
            amount_paid = request.POST.get("offers")
            new_payment = models.Payment.objects.create(
                user=request.user,
                reference=payment_reference,
                amount=amount_paid,
                transaction_date=datetime.now(),
                transaction_status="Pending",
                channel="telecel"
            )
            new_payment.save()
            phone_number = request.POST.get("phone")
            offer = request.POST.get("offers")
            if user.status == "User":
                bundle = models.VodaBundlePrice.objects.get(price=offer)
            elif user.status == "Agent":
                bundle = models.AgentVodaBundlePrice.objects.get(price=offer)
            elif user.status == "Super Agent":
                bundle = models.SuperAgentVodaBundlePrice.objects.get(price=offer)
            else:
                bundle = models.VodaBundlePrice.objects.get(price=offer)

            new_payment.transaction_details = {
                "amount": offer,
                "bundle": bundle,
                "phone_number": phone_number,
                "payment_reference": payment_reference
            }
            new_payment.save()
            return JsonResponse({"status": "Your transaction will be completed soon"})
        else:
            messages.info(request, "Invalid Channel")
            return redirect("home")

    # user = models.CustomUser.objects.get(id=request.user.id)
    # phone_num = user.phone
    # mtn_dict = {}
    #
    # if user.status == "Agent":
    #     mtn_offer = models.AgentMTNBundlePrice.objects.all()
    # else:
    #     mtn_offer = models.MTNBundlePrice.objects.all()
    # for offer in mtn_offer:
    #     mtn_dict[str(offer)] = offer.bundle_volume
    payment_method = models.AdminInfo.objects.filter().first().active_payment
    context = {'form': form,
               "ref": reference, "email": user_email, "wallet": 0 if user.wallet is None else user.wallet,
               'id': db_user_id, "payment_method": payment_method}
    return render(request, "layouts/services/voda.html", context=context)


@login_required(login_url='login')
def voda_pay_with_wallet(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        print(phone_number)
        print(amount)
        print(reference)
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge.'})

        if user.status == "User":
            bundle = models.VodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Agent":
            bundle = models.AgentVodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        elif user.status == "Super Agent":
            bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(amount)).bundle_volume
        else:
            bundle = models.VodaBundlePrice.objects.get(price=float(amount)).bundle_volume

        print(bundle)
        new_mtn_transaction = models.VodafoneTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{bundle}MB",
            reference=reference,
        )
        new_mtn_transaction.save()
        user.wallet -= float(amount)
        user.save()

        sms_message = f"Telecel order has been placed. {bundle}MB for {phone_number}. Reference: {reference}"
        response1 = requests.get(
            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0549914001&from=Noble Data&sms={sms_message}")
        print(response1.text)
        return JsonResponse({'status': "Your transaction will be completed shortly", 'icon': 'success'})
    return redirect('voda')


@login_required(login_url='login')
def voda_history(request):
    user_transactions = models.VodafoneTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "Vodafone Transactions"
    net = "voda"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def admin_voda_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.VodafoneTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/voda_admin.html", context=context)


@login_required(login_url='login')
def voda_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.VodafoneTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1317|sCtbw8U97Nwg10hVbZLBPXiJ8AUby7dyozZMjJpU',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your Vodafone transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH DATA',
            'message': sms_message
        }
        response1 = requests.get(
            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{txn.user.phone}&from=Noble Data&sms={sms_message}")
        print(response1.text)
        messages.success(request, f"Transaction Completed")
        return redirect('voda_admin')


@login_required(login_url='login')
def at_credit_history(request):
    user_transactions = models.ATMinuteTransaction.objects.filter(user=request.user).order_by(
        'transaction_date').reverse()
    header = "AT Minutes Transactions"
    net = "at_min"
    context = {'txns': user_transactions, "header": header, "net": net}
    return render(request, "layouts/history.html", context=context)


@login_required(login_url='login')
def admin_at_min_history(request):
    if request.user.is_staff and request.user.is_superuser:
        all_txns = models.ATMinuteTransaction.objects.filter().order_by('-transaction_date')[:1000]
        context = {'txns': all_txns}
        return render(request, "layouts/services/at_min_admin.html", context=context)


@login_required(login_url='login')
def at_min_mark_as_sent(request, pk):
    if request.user.is_staff and request.user.is_superuser:
        txn = models.ATMinuteTransaction.objects.filter(id=pk).first()
        print(txn)
        txn.transaction_status = "Completed"
        txn.save()
        sms_headers = {
            'Authorization': 'Bearer 1334|wroIm5YnQD6hlZzd8POtLDXxl4vQodCZNorATYGX',
            'Content-Type': 'application/json'
        }

        sms_url = 'https://webapp.usmsgh.com/api/sms/send'
        sms_message = f"Your AT Minutes transaction has been completed. {txn.bundle_number} has been credited with {txn.offer}.\nTransaction Reference: {txn.reference}"

        sms_body = {
            'recipient': f"233{txn.user.phone}",
            'sender_id': 'GH BAY',
            'message': sms_message
        }
        try:
            response1 = requests.get(
                f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{txn.user.phone}&from=Noble Data&sms={sms_message}")
            print(response1.text)
        except:
            messages.success(request, f"Transaction Completed")
            return redirect('at_min_admin')
        messages.success(request, f"Transaction Completed")
        return redirect('at_min_admin')


@login_required(login_url='login')
def pay_with_wallet_minutes(request):
    if request.method == "POST":
        admin = models.AdminInfo.objects.filter().first().phone_number
        user = models.CustomUser.objects.get(id=request.user.id)
        phone_number = request.POST.get("phone")
        amount = request.POST.get("amount")
        reference = request.POST.get("reference")
        if user.wallet is None:
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        elif user.wallet <= 0 or user.wallet < float(amount):
            return JsonResponse(
                {'status': f'Your wallet balance is low. Contact the admin to recharge. Admin Contact Info: 0{admin}'})
        print(phone_number)
        print(amount)
        print(reference)
        if user.status == "User":
            minutes = models.ATCreditPrice.objects.get(price=float(amount)).minutes
        elif user.status == "Agent":
            minutes = models.AgentATCreditPrice.objects.get(price=float(amount)).minutes
        elif user.status == "Super Agent":
            minutes = models.SuperAgentATCreditPrice.objects.get(price=float(amount)).minutes
        else:
            minutes = models.ATCreditPrice.objects.get(price=float(amount)).minutes

        new_transaction = models.ATMinuteTransaction.objects.create(
            user=request.user,
            bundle_number=phone_number,
            offer=f"{minutes} Minutes",
            reference=reference,
            transaction_status="Pending"
        )
        new_transaction.save()
        user.wallet -= float(amount)
        user.save()
        sms_message = f"AT Minutes order has been placed. {minutes} minutes for {phone_number}. Reference: {reference}"
        response1 = requests.get(
            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0549914001&from=Noble Data&sms={sms_message}")
        print(response1.text)

        return JsonResponse({'status': 'Transaction Received Successfully', 'icon': 'success'})
    return redirect('at_minutes')


@login_required(login_url='login')
def airtel_tigo_minutes(request):
    user = models.CustomUser.objects.get(id=request.user.id)
    status = user.status
    form = forms.ATCreditForm(status)
    reference = helper.ref_generator()
    db_user_id = request.user.id
    user_email = request.user.email
    # if request.method == "POST":
    #     form = forms.ATCreditForm(data=request.POST)
    #     payment_reference = request.POST.get("reference")
    #     amount_paid = request.POST.get("amount")
    #     new_payment = models.Payment.objects.create(
    #         user=request.user,
    #         reference=payment_reference,
    #         amount=amount_paid,
    #         transaction_date=datetime.now(),
    #         transaction_status="Completed"
    #     )
    #     new_payment.save()
    #     print("payment saved")
    #     print("form valid")
    #     phone_number = request.POST.get("phone")
    #     offer = request.POST.get("amount")
    #     print(offer)
    #     if user.status == "User":
    #         bundle = models.ATCreditPrice.objects.get(price=float(offer)).minutes
    #     elif user.status == "Agent":
    #         bundle = models.AgentATCreditPrice.objects.get(price=float(offer)).minutes
    #     else:
    #         bundle = models.ATCreditPrice.objects.get(price=float(offer)).minutes
    #     new_transaction = models.ATMinuteTransaction.objects.create(
    #         user=request.user,
    #         bundle_number=phone_number,
    #         offer=f"{bundle} Minutes",
    #         reference=payment_reference,
    #         transaction_status="Pending"
    #     )
    #     print("created")
    #     new_transaction.save()
    #
    #     print("===========================")
    #     print(phone_number)
    #     print(bundle)
    #     print("--------------------")
    #     # send_bundle_response = helper.send_bundle(request.user, phone_number, bundle, payment_reference)
    #     # print("********************")
    #     # data = send_bundle_response.json()
    #     #
    #     # print(data)
    #     return JsonResponse({'status': 'Transaction Completed Successfully', 'icon': 'success'})
    user = models.CustomUser.objects.get(id=request.user.id)
    context = {"form": form, "ref": reference, 'id': db_user_id, "email": user_email,
               "wallet": 0 if user.wallet is None else user.wallet}
    return render(request, "layouts/services/at_credit.html", context=context)


def anmgw_payment_page(request, reference):
    form = forms.AppsNMobileForm()
    user = models.CustomUser.objects.get(id=request.user.id)
    if request.method == "POST":
        form = forms.AppsNMobileForm(request.POST)
        if form.is_valid():
            phone_number = form.cleaned_data['phone_number']
            mobile_network = form.cleaned_data['mobile_network']
            reference = reference

            print(phone_number)
            print(mobile_network)
            print(reference)

            payment = models.Payment.objects.get(reference=reference)
            amount = payment.amount
            callback_url = "https://www.nobledatagh.com/callback_url"
            service_id = int(config("ANM_SERVICE_ID"))
            print(service_id)

            send_payment_response = send_payment_request(user, amount, f"233{phone_number}", reference, mobile_network,
                                                         callback_url, service_id)
            print(send_payment_response)
            if send_payment_response:
                response_code = send_payment_response['resp_code']
                if response_code == "015":
                    payment.transaction_status = "Processing"
                    payment.save()
                    messages.info(request, "You will receive a mobile money prompt soon.")
                    return JsonResponse(data={"status": "Success"})
                else:
                    messages.info(request, "Something went wrong. Try again")
                    return JsonResponse(data={"status": "Failed"})
            else:
                messages.error(request, "Payment Processing Failed")
                return redirect('anmgw_payment_page', reference=reference)
        else:
            messages.error(request, "Invalid Submission")
            return redirect('anmgw_payment_page', reference=reference)
    if not models.Payment.objects.filter(reference=reference).exists():
        messages.info(request, "Link broken.")
        return redirect('home')
    try:
        if models.Payment.objects.filter(reference=reference, transaction_status="Completed").exists():
            messages.info(request, "Payment already completed.")
            return redirect('home')
    except models.Payment.DoesNotExist:
        pass

    payment = models.Payment.objects.get(reference=reference)
    amount = payment.amount
    channel = payment.channel
    reference = reference
    payment_method = models.AdminInfo.objects.filter().first().active_payment  # Get the active payment method

    context = {
        'form': form,
        'amount': amount,
        'channel': channel,
        'reference': reference,
        'payment_method': payment_method  # Pass the payment method to the template
    }
    return render(request, "layouts/payments/anmgw.html", context=context)


def check_payment_status(request, reference):
    payment = Payment.objects.get(reference=reference)
    return JsonResponse({'status': payment.transaction_status})


def payment_receipt(request, reference):
    payment = Payment.objects.get(reference=reference)
    context = {
        'payment': payment
    }
    return render(request, 'layouts/payments/receipt.html', context)


def generate_hmac_signature(payload, secret_key):
    message = json.dumps(payload).encode('utf-8')
    secret = secret_key.encode('utf-8')
    signature = hmac.new(secret, message, hashlib.sha256).hexdigest()
    return signature


def send_payment_request(user, amount, phone_number, reference, nw, callback_url, service_id):
    # Generate a unique exttrid
    exttrid = generate_unique_exttrid()

    # Create a new payment record
    payment = Payment(
        user=user,
        reference=exttrid,
        amount=amount,
        channel=nw,
        payment_description=reference,
        transaction_status="Unfinished",
        transaction_date=timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    )
    payment.save()

    data = {
        "customer_number": phone_number,  # Assuming user has a phone_number field
        "amount": f"{amount:.2f}",
        "exttrid": reference,
        "reference": "Payment to Noble Data",
        "nw": nw,
        "trans_type": "CTM",  # Collect money from customer's mobile money wallet
        "callback_url": callback_url,
        "service_id": service_id,
        "ts": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),  # Current timestamp in UTC
        "nickname": "Noble Data"  # Optional, set as "Noble Data"
    }

    # Prepare API request data
    url = "https://orchard-api.anmgw.com/sendRequest"
    client_id = config("ANM_CLIENT_TOKEN")
    client_secret = config("ANM_SECRET_TOKEN")
    signature = generate_hmac_signature(data, client_secret)
    authorization_header = f"{client_id}:{signature}"
    print(client_id)
    print(client_secret)

    headers = {
        'Authorization': authorization_header,  # Replace with actual CLIENT_ID and SIGNATURE
        'Content-Type': 'application/json'
    }

    response = requests.post(url, headers=headers, json=data)
    return response.json()


def anmgw_callback(request):
    if request.method == "POST":
        payload = request.body.decode('utf-8')
        print("aNM payment Info: ", payload)
        json_payload = json.loads(payload)
        print(json_payload)

        # {
        #     "trans_id": "98765432100",
        #     "trans_ref": 4243846303,
        #     "trans_status": "000/01",
        #     "message": "SUCCESS"
        # }

        reference = json_payload['trans_ref']
        txn_status = json_payload['trans_status']
        transaction_id = json_payload['trans_id']

        if txn_status == '000/01':
            print("success")
            transaction_saved = models.Payment.objects.get(reference=reference, transaction_status="Unfinished")
            transaction_saved.transaction_status = "Completed"
            transaction_saved.payment_description = transaction_id
            transaction_saved.save()
            transaction_details = transaction_saved.transaction_details
            real_amount = transaction_details['amount']
            receiver = transaction_details["phone_number"]
            transaction_channel = transaction_saved.channel
            user = transaction_saved.user
            print(transaction_details, transaction_channel)

            if transaction_channel == "ishare":
                if user.status == "User":
                    bundle = models.IshareBundlePrice.objects.get(price=real_amount)
                elif user.status == "Agent":
                    bundle = models.AgentIshareBundlePrice.objects.get(price=real_amount)
                elif user.status == "Super Agent":
                    bundle = models.SuperAgentIshareBundlePrice.objects.get(price=real_amount)
                else:
                    bundle = models.VodaBundlePrice.objects.get(price=real_amount)
                if models.IShareBundleTransaction.objects.filter(reference=reference, offer=f"{bundle}MB",
                                                                 transaction_status="Completed").exists():
                    return HttpResponse(status=200)
                new_transaction = models.IShareBundleTransaction.objects.create(
                    user=user,
                    bundle_number=receiver,
                    offer=f"{bundle}MB",
                    reference=reference,
                    transaction_status="Pending"
                )
                new_transaction.save()
                send_bundle_response = helper.send_bundle(user, receiver, bundle, reference)
                data = send_bundle_response.json()

                print(data)

                sms_headers = {
                    'Authorization': 'Bearer 1050|VDqcCUHwCBEbjcMk32cbdOhCFlavpDhy6vfgM4jU',
                    'Content-Type': 'application/json'
                }

                sms_url = 'https://webapp.usmsgh.com/api/sms/send'

                if send_bundle_response.status_code == 200:
                    if data["code"] == "0000":
                        transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                            reference=reference)
                        print("got here")
                        print(transaction_to_be_updated.transaction_status)
                        transaction_to_be_updated.transaction_status = "Completed"
                        transaction_to_be_updated.save()

                        num_without_0 = receiver[1:]

                        receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                        sms_message = f"Hello @{user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {receiver}.\nReference: {reference}\nThank you for using Noble Data GH."

                        response1 = requests.get(
                            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0{user.phone}&from=Noble Data&sms={sms_message}")
                        print(response1.text)

                        response2 = requests.get(
                            f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to={num_without_0}&from=Noble Data&sms={receiver_message}")
                        print(response2.text)

                        bundle = models.IshareBundlePrice.objects.get(price=float(
                            real_amount)) if user.status == "User" else models.AgentIshareBundlePrice.objects.get(
                            price=float(real_amount))

                        purchase_price = bundle.purchase_price

                        profit = float(real_amount) - float(purchase_price)

                        new_profit_instance = models.ProfitInstance.objects.create(
                            selling_price_total=real_amount,
                            purchase_price_total=purchase_price,
                            profit=profit,
                            channel="AT",  # Set your channel here based on user status
                        )
                        new_profit_instance.save()

                        print(user.phone)
                        print("***********")
                        receiver_message = f"Your bundle purchase has been completed successfully. {bundle}MB has been credited to you by {user.phone}.\nReference: {reference}\n"
                        sms_message = f"Hello @{user.username}. Your bundle purchase has been completed successfully. {bundle}MB has been credited to {receiver}.\nReference: {reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"

                        print(num_without_0)
                        # receiver_body = {
                        #     'recipient': f"233{num_without_0}",
                        #     'sender_id': 'Noble Data',
                        #     'message': receiver_message
                        # }
                        #
                        # response = requests.request('POST', url=sms_url, params=receiver_body, headers=sms_headers)
                        # print(response.text)
                        #
                        # sms_body = {
                        #     'recipient': f"233{user.phone}",
                        #     'sender_id': 'Noble Data',
                        #     'message': sms_message
                        # }
                        #
                        # response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        #
                        # print(response.text)

                        return HttpResponse(status=200)
                    else:
                        transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                            reference=reference)
                        transaction_to_be_updated.transaction_status = "Failed"
                        new_transaction.save()
                        sms_message = f"Hello @{user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {receiver}.\nReference: {reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"

                        sms_body = {
                            'recipient': f"233{user.phone}",
                            'sender_id': 'Noble Data',
                            'message': sms_message
                        }
                        response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)
                        print(response.text)
                        return HttpResponse(status=500)
                else:
                    transaction_to_be_updated = models.IShareBundleTransaction.objects.get(
                        reference=reference)
                    transaction_to_be_updated.transaction_status = "Failed"
                    new_transaction.save()
                    sms_message = f"Hello @{request.user.username}. Something went wrong with your transaction. Contact us for enquiries.\nBundle: {bundle}MB\nPhone Number: {receiver}.\nReference: {reference}\nThank you for using Noble Data GH.\n\nThe Noble Data GH"

                    sms_body = {
                        'recipient': f'233{user.phone}',
                        'sender_id': 'Noble Data',
                        'message': sms_message
                    }

                    response = requests.request('POST', url=sms_url, params=sms_body, headers=sms_headers)

                    print(response.text)
                    return HttpResponse(status=500)
            elif transaction_channel == "mtn":
                if user.status == "User":
                    bundle = models.MTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                elif user.status == "Agent":
                    bundle = models.AgentMTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                elif user.status == "Super Agent":
                    bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                else:
                    bundle = models.MTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                new_mtn_transaction = models.MTNTransaction.objects.create(
                    user=user,
                    bundle_number=receiver,
                    offer=f"{bundle}MB",
                    reference=reference,
                )
                new_mtn_transaction.save()

                print(receiver)

                if user.status == "User":
                    bundle = models.MTNBundlePrice.objects.get(price=float(real_amount))
                elif user.status == "Agent":
                    bundle = models.AgentMTNBundlePrice.objects.get(price=float(real_amount))
                elif user.status == "Super Agent":
                    bundle = models.SuperAgentMTNBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                else:
                    bundle = models.MTNBundlePrice.objects.get(price=float(real_amount))

                purchase_price = bundle.purchase_price

                profit = float(real_amount) - float(purchase_price)

                new_profit_instance = models.ProfitInstance.objects.create(
                    selling_price_total=real_amount,
                    purchase_price_total=purchase_price,
                    profit=profit,
                    channel="MTN",  # Set your channel here based on user status
                )
                new_profit_instance.save()

                return HttpResponse(status=200)
            elif transaction_channel == "telecel":
                if user.status == "User":
                    bundle = models.VodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                elif user.status == "Agent":
                    bundle = models.AgentVodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                elif user.status == "Super Agent":
                    bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume
                else:
                    bundle = models.VodaBundlePrice.objects.get(price=float(real_amount)).bundle_volume

                new_voda_transaction = models.VodafoneTransaction.objects.create(
                    user=user,
                    bundle_number=receiver,
                    offer=f"{bundle}MB",
                    reference=reference,
                )
                new_voda_transaction.save()

                sms_message = f"Telecel order has been placed. {bundle}MB for {receiver}. Reference: {reference}"
                response1 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0549914001&from=Noble Data&sms={sms_message}")
                print(response1.text)

                print(receiver)

                if user.status == "User":
                    bundle = models.VodaBundlePrice.objects.get(price=float(real_amount))
                elif user.status == "Agent":
                    bundle = models.AgentVodaBundlePrice.objects.get(price=float(real_amount))
                elif user.status == "Super Agent":
                    bundle = models.SuperAgentVodaBundlePrice.objects.get(price=float(real_amount))
                else:
                    bundle = models.VodaBundlePrice.objects.get(price=float(real_amount))

                purchase_price = bundle.purchase_price

                profit = float(real_amount) - float(purchase_price)

                new_profit_instance = models.ProfitInstance.objects.create(
                    selling_price_total=real_amount,
                    purchase_price_total=purchase_price,
                    profit=profit,
                    channel="Voda",  # Set your channel here based on user status
                )
                new_profit_instance.save()

                return HttpResponse(status=200)
            elif transaction_channel == "at_min":
                if user.status == "User":
                    minutes = models.ATCreditPrice.objects.get(price=float(real_amount)).minutes
                elif user.status == "Agent":
                    minutes = models.AgentATCreditPrice.objects.get(price=float(real_amount)).minutes
                elif user.status == "Super Agent":
                    minutes = models.SuperAgentATCreditPrice.objects.get(price=float(real_amount)).minutes
                else:
                    minutes = models.ATCreditPrice.objects.get(price=float(real_amount)).minutes

                print(receiver)

                new_mtn_transaction = models.ATMinuteTransaction.objects.create(
                    user=user,
                    bundle_number=receiver,
                    offer=f"{minutes} Minutes",
                    reference=reference,
                )
                new_mtn_transaction.save()
                sms_message = f"AT Minutes order has been placed. {minutes} minutes for {receiver}. Reference: {reference}"
                response1 = requests.get(
                    f"https://sms.arkesel.com/sms/api?action=send-sms&api_key=Ojd0Uk5BVmE3SUpna2lRS3o=&to=0549914001&from=Noble Data&sms={sms_message}")
                print(response1.text)
                return HttpResponse(status=200)
            elif transaction_channel == "topup":
                amount = real_amount

                user.wallet += float(amount)
                user.save()

                new_topup = models.TopUpRequest.objects.create(
                    user=user,
                    reference=reference,
                    amount=amount,
                    status=True,
                )
                new_topup.save()

                new_wallet_transaction = models.WalletTransaction.objects.create(
                    user=user,
                    transaction_type="Credit",
                    transaction_amount=float(amount),
                    transaction_use="Top up (appsNMobile)",
                    new_balance=user.wallet
                )
                new_wallet_transaction.save()
                return JsonResponse({'status': "Wallet Credited"}, status=200)


